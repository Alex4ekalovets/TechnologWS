import json
import logging
import math
import os
import shutil
import sys
import threading
import time
from copy import copy

import requests

from PyQt6.QtCore import Qt, QAbstractTableModel, QSize, pyqtSignal, QRegularExpression
from PyQt6.QtGui import QIcon, QAction, QRegularExpressionValidator, QColor
from PyQt6.QtWidgets import (
    QWidget, QApplication,
    QGridLayout, QLabel, QTableView, QMainWindow,
    QFileDialog, QStatusBar, QToolBar, QComboBox, QDialog, QPushButton, QSplitter, QLineEdit, QCheckBox, QMessageBox,
)
import pandas as pd
import numpy as np
from openpyxl.reader.excel import load_workbook

from cdw_reader import get_specification
from graph import draw_graph
from parse_operations import get_all_operations, COLUMNS


def df_handler(df, start_id=0):
    columns_count = df.shape[1] if df.shape[1] < 20 else 20
    filtered_data = df[df.iloc[:, 2].notna() & df.iloc[:, 13].astype(str).str.isdigit()].iloc[:, 0:columns_count]
    filtered_data.iloc[:, 0:2] = filtered_data.iloc[:, 0:2].ffill()
    filtered_data.columns = COLUMNS[:columns_count]
    filtered_data = filtered_data.reindex(columns=COLUMNS)
    filtered_data['id'] = range(start_id + 1, len(filtered_data) + start_id + 1)
    filtered_data[COLUMNS[21]] = np.empty((len(filtered_data), 0)).tolist()
    filtered_data[COLUMNS[22]] = np.empty((len(filtered_data), 0)).tolist()
    return filtered_data


class PandasModel(QAbstractTableModel):
    def __init__(self, data, last_id, planned=False, immutables=()):
        super().__init__()
        self.df = data
        self.last_id = last_id
        self.planned = planned
        self.colors = dict()
        self.immutables = immutables

    def rowCount(self, index):
        return self.df.shape[0]

    def columnCount(self, parent=None):
        return self.df.shape[1]

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        try:
            if index.isValid():
                if role == Qt.ItemDataRole.BackgroundRole:
                    color = self.colors.get((index.row(), index.column()))
                    if color is not None:
                        return color
                if role == Qt.ItemDataRole.DisplayRole or role == Qt.ItemDataRole.EditRole:
                    value = self.df.iloc[index.row(), index.column()]
                    if str(value) == 'nan' or str(value) == 'None':
                        value = f''
                    elif isinstance(value, float):
                        value = f'{value:.2f}'
                    else:
                        value = f'{value}'
                    return value
        except Exception as ex:
            logging.exception(ex)

    def setData(self, index, value, role):
        try:
            def val(i):
                return float(self.df.iloc[index.row(), i])

            if role == Qt.ItemDataRole.EditRole:
                if index.column() not in [20, 21, 22] + list(range(6, 15)):
                    if value.strip() == '':
                        value = None
                    self.change_color(index.row(), index.column(), QColor("white"))
                    self.df.iloc[index.row(), index.column()] = value
                elif index.column() == 6:
                    if float(value) == 0:
                        logging.error('Деление на 0')
                        return False
                    self.df.iloc[index.row(), 6] = value
                    self.df.iloc[index.row(), 9] = val(7) * val(8) / val(6)
                    self.df.iloc[index.row(), 11] = val(9) * val(10)
                    self.df.iloc[index.row(), 12] = val(11) * val(13)
                    self.df.iloc[index.row(), 14] = val(9) * val(10) / val(8)
                elif index.column() == 7:
                    self.df.iloc[index.row(), 7] = value
                    self.df.iloc[index.row(), 6] = val(8) * val(7) / val(9)
                elif index.column() == 8:
                    if float(value) == 0:
                        logging.error('Деление на 0')
                        return False
                    self.df.iloc[index.row(), 8] = value
                    self.df.iloc[index.row(), 6] = val(8) * val(7) / val(9)
                    self.df.iloc[index.row(), 14] = val(9) * val(10) / val(8)
                elif index.column() == 9:
                    if float(value) == 0:
                        logging.error('Деление на 0')
                        return False
                    self.df.iloc[index.row(), 9] = value
                    self.df.iloc[index.row(), 6] = val(8) * val(7) / val(9)
                    self.df.iloc[index.row(), 11] = val(9) * val(10)
                    self.df.iloc[index.row(), 12] = val(11) * val(13)
                    self.df.iloc[index.row(), 14] = val(9) * val(10) / val(8)
                elif index.column() == 10:
                    self.df.iloc[index.row(), 10] = value
                    self.df.iloc[index.row(), 11] = val(9) * val(10)
                    self.df.iloc[index.row(), 12] = val(11) * val(13)
                    self.df.iloc[index.row(), 14] = val(9) * val(10) / val(8)
                elif index.column() == 13:
                    self.df.iloc[index.row(), 13] = value
                    self.df.iloc[index.row(), 12] = val(11) * val(13)
                    self.df.iloc[index.row(), 14] = val(9) * val(10) / val(8)
                # self.df.to_json(r'data\abc.json', orient='records')
                return True
        except Exception as ex:
            logging.exception(ex)
        finally:
            return False

    def headerData(self, section, orientation, role):
        # section is the index of the column/row.
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return str(self.df.columns[section])

            if orientation == Qt.Orientation.Vertical:
                return str(self.df.index[section] + 1)

    def flags(self, index):
        try:
            if self.df.shape[1] > 19 and self.df.iloc[index.row(), 20] in self.immutables and index.column() not in [19,
                                                                                                                     21,
                                                                                                                     22]:
                return Qt.ItemFlag.ItemIsSelectable
            elif self.planned and index.column() in [200] or index.column() in [11, 12, 14]:
                return Qt.ItemFlag.ItemIsSelectable
            else:
                return Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable
        except Exception as ex:
            logging.exception(ex)

    def change_color(self, row, column, color):
        ix = self.index(row, column)
        self.colors[(row, column)] = color
        self.dataChanged.emit(ix, ix, (Qt.ItemDataRole.BackgroundRole,))


class Window(QMainWindow):
    def __init__(self):
        super().__init__(parent=None)
        self.setWindowTitle("Рабочее место технолога")
        self.setMinimumSize(1200, 600)
        self.setWindowIcon(QIcon("img/icon.svg"))

        self._create_actions()
        self._create_menu()
        self._create_toolbar()
        self._connect_actions()
        self._create_statusbar()
        self._open_main_window()

    def _create_menu(self):
        menu_bar = self.menuBar()

        file_menu = menu_bar.addMenu("&Файл")
        edit_menu = menu_bar.addMenu("&Правка")
        help_menu = menu_bar.addMenu("&Справка")

        file_menu.addAction(self.open_action)
        file_menu.addSeparator()
        file_menu.addAction(QIcon("img/close.svg"), "&Выход", self.close)

        edit_menu.addAction(self.undo_action)
        edit_menu.addAction(self.redo_action)

        help_menu.addAction("&О программе")

    def _create_toolbar(self):
        tools = QToolBar("Навигация")
        tools.setIconSize(QSize(35, 35))
        tools.addAction(self.open_main_window_action)
        tools.addAction(self.open_plasma_window_action)
        tools.addSeparator()
        tools.addAction(self.close_app_action)
        self.addToolBar(Qt.ToolBarArea.LeftToolBarArea, tools)

    def _create_statusbar(self):
        status = QStatusBar()
        self.setStatusBar(status)

    def _create_actions(self):
        self.open_main_window_action = QAction(QIcon("img/process.png"), "Технологический процесс", self)
        self.open_plasma_window_action = QAction(QIcon("img/plasma.png"), "Плазма", self)
        self.close_app_action = QAction(QIcon("img/exit.svg"), "Выход", self)
        self.open_action = QAction(QIcon("img/open.svg"), "&Открыть", self)
        self.undo_action = QAction(QIcon("img/undo.svg"), "&Отменить", self)
        self.redo_action = QAction(QIcon("img/redo.svg"), "&Повторить", self)

    def _remove_previous_widget_toolbar(self):
        if hasattr(self, 'widget_tools'):
            self.removeToolBar(self.widget_tools)

    def _open_main_window(self):
        if hasattr(self, 'central_widget') and isinstance(self.central_widget, WorkspaceWidget):
            return
        self._remove_previous_widget_toolbar()
        self.central_widget = WorkspaceWidget(self)
        self.setCentralWidget(self.central_widget)

    def _connect_actions(self):
        self.open_main_window_action.triggered.connect(self._open_main_window)
        self.close_app_action.triggered.connect(self.close)


class OrderModelSelect(QComboBox):
    clicked = pyqtSignal()

    def showPopup(self):
        self.clicked.emit()
        super(OrderModelSelect, self).showPopup()


class WorkspaceWidget(QWidget):
    def __init__(self, main_window):
        super().__init__()
        self.xlsx_file_name = ''
        self.main_window = main_window
        self.layout = QGridLayout()
        self.layout.setSpacing(10)
        self._create_actions()
        self._create_main_widget_toolbar()
        self.displayed_columns = COLUMNS
        self.search_data = None
        self.orders_models = []
        self.is_planned = False
        self.process_is_upload = False
        self.has_fio_doers = []
        self.is_on_change = True

        self.setLayout(self.layout)
        self.table_process = QTableView()
        self.table_process.setDisabled(True)
        self.table_process.doubleClicked.connect(self.double_click_process_table)
        self._create_process_table_context_menu()

        self.table_details = QTableView()
        self._create_details_table_context_menu()
        self._connect_actions()

        self.initial_data = pd.DataFrame(
            [],
            columns=COLUMNS
        )
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.set_table_data(self.table_process, self.initial_data)
        self.splitter.addWidget(self.table_process)
        self.splitter.addWidget(self.table_details)

        self.searchbar = QLineEdit()
        self.searchbar.textChanged.connect(self.search)

        order_model_label = QLabel()
        order_model_label.setText('Выберите заказ-модель:')
        self.order_model_select = OrderModelSelect()
        self.order_model_select.setMinimumWidth(160)
        self.order_model_select.addItem("Не выбрано")
        self.order_model_select.clicked.connect(self.on_order_model_click)
        self.order_model_select.activated.connect(self.on_order_model_select)

        model_label = QLabel()
        model_label.setText('Модель:')
        self.model_edit = QLineEdit()
        self.model_edit.setMinimumWidth(160)
        model_reg_ex = QRegularExpression(r"^[\-A-Za-z0-9]+$")
        model_validator = QRegularExpressionValidator(model_reg_ex, self.model_edit)
        self.model_edit.setValidator(model_validator)
        self.model_edit.setToolTip(
            "Имя модели может содержать только цифры и буквы латинского алфавита и знак '-' тире"
        )
        self.model_edit.setDisabled(True)

        self.process_status_label = QLabel()
        self.process_status_label.setText('Статус:')

        self.table_details.hide()
        columns_select_button = QPushButton(text="", parent=self)
        columns_select_button.setIcon(QIcon("img/select-columns.svg"))
        columns_select_button.clicked.connect(self.open_select_columns_dialog)

        self.change_button = QPushButton(text="Корректировка", parent=self)
        self.change_button.hide()
        self.change_button.clicked.connect(self.take_on_change)

        self.layout.addWidget(order_model_label, 0, 0, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.order_model_select, 0, 1, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(model_label, 0, 2, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.model_edit, 0, 3, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.process_status_label, 0, 5, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.change_button, 0, 4, 1, 1, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(columns_select_button, 1, 1, 1, 9, Qt.AlignmentFlag.AlignRight)
        self.layout.addWidget(self.searchbar, 1, 0, 1, 2, Qt.AlignmentFlag.AlignLeft)
        self.layout.addWidget(self.splitter, 2, 0, 1, 10)

    def _create_main_widget_toolbar(self):
        self.widget_tools = QToolBar("Инструменты")
        self.widget_tools.setIconSize(QSize(35, 35))
        self.widget_tools.addAction(self.save_action)
        self.widget_tools.addAction(self.upload_draws_action)
        self.widget_tools.addAction(self.open_xlsx_action)
        self.widget_tools.addAction(self.upload_process_action)
        self.widget_tools.addAction(self.draw_graph_action)
        self.widget_tools.addAction(self.save_to_xlsx_action)
        self.main_window.addToolBar(Qt.ToolBarArea.TopToolBarArea, self.widget_tools)

    def _create_actions(self):
        self.upload_draws_action = QAction(QIcon("img/upload.svg"), "Загрузить детали из CDW", self)
        self.open_xlsx_action = QAction(QIcon("img/xlsx.svg"), "&Загрузить техпроцесс из XLSX", self)
        self.add_row_up_action = QAction(QIcon("img/add-row-up.png"), "&Добавить строку выше", self)
        self.add_row_down_action = QAction(QIcon("img/add-row-down.png"), "&Добавить строку ниже", self)
        self.delete_row_action = QAction(QIcon("img/delete-row.png"), "&Удалить строку", self)
        self.add_detail_to_process_action = QAction(QIcon("img/add-operation.svg"), "&Создать операцию для детали",
                                                    self)
        self.add_operation_action = QAction(QIcon("img/add-operation.svg"), "&Добавить операцию", self)
        self.upload_process_action = QAction(QIcon("img/upload-process.svg"), "&Загрузить техпроцесс", self)
        self.draw_graph_action = QAction(QIcon("img/graph.svg"), "&Нарисовать техпроцесс", self)
        self.save_action = QAction(QIcon("img/save.svg"), "&Сохранить изменения", self)
        self.save_to_xlsx_action = QAction(QIcon("img/to-xlsx.png"), "&Сохранить в XLSX", self)

    def _connect_actions(self):
        self.upload_draws_action.triggered.connect(self.open_cdw_dialog)
        self.open_xlsx_action.triggered.connect(self.open_xlsx_dialog)
        self.add_row_up_action.triggered.connect(self.add_row_up)
        self.add_row_down_action.triggered.connect(self.add_row_down)
        self.delete_row_action.triggered.connect(self.delete_row)
        self.add_detail_to_process_action.triggered.connect(self.add_operation_for_detail)
        self.add_operation_action.triggered.connect(self.add_operation)
        self.upload_process_action.triggered.connect(self.upload_process)
        self.draw_graph_action.triggered.connect(self.draw_graph)
        self.save_action.triggered.connect(self.save_process_to_file)
        self.save_to_xlsx_action.triggered.connect(self.save_to_xlsx)

    def open_xlsx_dialog(self):
        self.xlsx_file_name = QFileDialog.getOpenFileName(
            self,
            caption="Выберите файл",
            directory=r"D:\\",
            filter="Excel Files (*.xlsx);;",
        )[0]
        if self.xlsx_file_name:
            xlsx_file = pd.ExcelFile(self.xlsx_file_name)
            sheet_names = xlsx_file.sheet_names
            dialog = SelectXlsxSheetWindow(self, sheet_names)
            dialog.exec()

            if dialog.selected_sheet:
                try:
                    if self.has_fio_doers:
                        data = df_handler(
                            pd.read_excel(self.xlsx_file_name, dialog.selected_sheet),
                            start_id=self.table_process.model().last_id
                        )
                        data = pd.concat((self.table_process.model().df, data))
                        self.table_process.model().last_id = data[COLUMNS[20]].max() + 1
                    else:
                        data = df_handler(pd.read_excel(self.xlsx_file_name, dialog.selected_sheet))
                        self.table_process.setModel(None)
                    self.set_table_data(self.table_process, data)
                except Exception as ex:
                    logging.exception(ex)
            else:
                self.set_table_data(self.table_process, self.initial_data)

    def open_cdw_dialog(self):
        file_names = QFileDialog.getOpenFileNames(
            self,
            caption="Выберите файлы",
            directory=r"D:\\",
            filter="CDW Files (*.cdw);;",
        )[0]
        thread = threading.Thread(target=self.load_cdw, args=(file_names,))
        thread.start()

    def set_table_data(self, table, data):

        logging.debug('Начало построения таблицы')
        data = data.sort_index()
        data.index = range(0, len(data))
        if table.model():
            last_id = table.model().last_id
        else:
            last_id = len(data)
        is_planned_process = table == self.table_process and self.is_planned
        if not self.is_on_change:
            immutables = data[COLUMNS[20]].tolist()
        else:
            immutables = self.has_fio_doers
        model = PandasModel(data, last_id, is_planned_process, immutables=immutables)
        try:
            table.setModel(model)
            logging.debug('Завершение построения таблицы')
        except Exception as ex:
            logging.exception(ex)

    def load_cdw(self, file_names):
        if file_names:
            self.main_window.statusBar().showMessage("Загрузка файлов...")
            json_file = get_specification(file_names)
            data = pd.read_json(json_file, orient='records')
            os.remove(json_file)
            if data.empty:
                data = pd.DataFrame()
            self.set_table_data(self.table_details, data)
            self.main_window.statusBar().clearMessage()
            self.table_details.show()

    def send_tech_data(self, data):
        self.main_window.statusBar().showMessage("Отправка техпроцесса...")
        try:
            url = "http://127.0.0.1:8000/tehnolog/tech_data"
            payload = json.dumps(data)
            response = requests.post(url, data=payload)
            data = json.loads(response.content)
            logging.debug(data['message'])
            if response.status_code == 200:
                self.save_process_to_file(upload=True)
                self.save_process_to_file()
                order = self.order_model_select.currentText().split('_')[0]
                self.on_order_model_click()
                self.order_model_select.setCurrentText(f'{order}_{self.model_edit.text()}')
                self.on_order_model_select(self.order_model_select.currentIndex())
            self.main_window.statusBar().showMessage(data['message'])
        except requests.exceptions.ConnectionError as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("⛔Нет связи с сервером!")
        except Exception as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("Ошибка отправки техпроцесса!")
        # finally:
        #     time.sleep(5)
        #     self.main_window.statusBar().clearMessage()

    def upload_process(self):
        try:
            data = self.table_process.model().df
            if self.is_valid(data):
                shift_tasks = {
                    'model_order_query': f'{self.order_model_select.currentText()}',
                    'order': self.order_model_select.currentText().split('_')[0],
                    'model_name': f'{self.model_edit.text()}',
                    'shift_tasks': [],
                }
                for i in range(len(data)):
                    shift_task = {
                        'model_order_query': f'{self.order_model_select.currentText()}',
                        'order': self.order_model_select.currentText().split('_')[0],
                        'model_name': f'{self.model_edit.text()}',
                        'op_number': f'{data.iloc[i, 0]}',
                        'op_name': f'{data.iloc[i, 1]}',
                        'ws_name': f'{data.iloc[i, 2]}',
                        'op_name_full': f'{data.iloc[i, 1]}-{data.iloc[i, 2]}',
                        'ws_number': f'{data.iloc[i, 3]}',
                        'norm_tech': float(data.iloc[i, 11]),
                        'draw_filename': f'{data.iloc[i, 19]}',
                        'tech_id': int(data.iloc[i, 20]),
                        'next_ids': [int(x) for x in data.iloc[i, 21]],
                        'prev_ids': [int(x) for x in data.iloc[i, 22]],
                    }
                    shift_tasks['shift_tasks'].append(shift_task)
                self.send_tech_data(shift_tasks)
                # thread = threading.Thread(target=self.send_tech_data, args=(shift_tasks,))
                # thread.start()
        except Exception as ex:
            logging.exception(ex)

    def is_valid(self, data):
        try:
            message = "⛔"
            terminals = ['11', '12']
            is_valid = True

            # Проверка 1 - Уникальность id
            is_valid = data[COLUMNS[20]].is_unique
            logging.debug(f"ID уникальны: {data[COLUMNS[20]].is_unique}")

            # Проверка 2 - Все обязательные поля заполнены (столбцы 0, 1, 3, 11)
            empty_cells = data[[COLUMNS[0], COLUMNS[1], COLUMNS[3], COLUMNS[11]]].stack(dropna=False)
            cells = [list(x) for x in empty_cells.index[empty_cells.isna() | empty_cells.isin(['', ])]]
            is_valid = is_valid and not cells
            if cells:
                for cell in cells:
                    row = cell[0]
                    column = COLUMNS.index(cell[1])
                    self.table_process.model().change_color(row, column, QColor("red"))
                message += "Заполните обязательные поля, выделенные красным! "
                is_valid = False

            # Проверка 3 - Номера терминалов указаны верно
            cells = [x for x in data.index[~data[COLUMNS[3]].astype(str).isin(terminals)]]
            if cells:
                for cell in cells:
                    row = cell
                    column = 3
                    self.table_process.model().change_color(row, column, QColor("red"))
                message += f"Номера терминалов должны быть из списка {terminals} "
                is_valid = False

            # Проверка 4 - Заполнена модель
            if self.model_edit.text() == '':
                message += 'Заполните модель! '
                is_valid = False

            # Проверка 5 - Выбран заказ-модель
            if self.order_model_select.currentText() == 'Не выбрано':
                message += 'Выберите заказ-модель! '
                is_valid = False

            self.main_window.statusBar().showMessage(message)
            return is_valid
        except Exception as ex:
            logging.exception(ex)

    def _create_process_table_context_menu(self):
        for action in self.table_process.actions():
            self.table_process.removeAction(action)
        self.table_process.setContextMenuPolicy(Qt.ContextMenuPolicy.ActionsContextMenu)
        if self.is_on_change:
            self.table_process.addAction(self.add_operation_action)
            self.table_process.addAction(self.add_row_up_action)
            self.table_process.addAction(self.add_row_down_action)
            self.table_process.addAction(self.delete_row_action)

    def _create_details_table_context_menu(self):
        self.table_details.setContextMenuPolicy(Qt.ContextMenuPolicy.ActionsContextMenu)
        self.table_details.addAction(self.add_detail_to_process_action)
        self.table_details.addAction(self.add_row_up_action)
        self.table_details.addAction(self.add_row_down_action)
        self.table_details.addAction(self.delete_row_action)

    def add_empty_row(self, table, name):
        df = table.model().df
        df.loc[name] = ''
        if table == self.table_process:
            table.model().last_id += 1
            df.loc[name, 'id'] = table.model().last_id
            df.at[name, COLUMNS[21]] = []
            df.at[name, COLUMNS[22]] = []
        return df

    def add_row(self, coeff):
        try:
            table = self.focusWidget()
            index = table.currentIndex().row()
            data = table.model().df
            if index == -1 and len(data) == 0:
                data = self.add_empty_row(table, 0)
            else:
                current_row = data.iloc[index].name
                new_name = current_row + coeff
                data = self.add_empty_row(table, new_name)
            self.set_table_data(table, data)
        except Exception as ex:
            logging.exception(ex)

    def add_row_up(self):
        self.add_row(coeff=-0.5)

    def add_row_down(self):
        self.add_row(coeff=0.5)

    def delete_row(self):
        try:
            table = self.focusWidget()
            indexes = set(map(lambda x: x.row(), table.selectedIndexes()))
            data = table.model().df
            row_names = []
            for index in indexes:
                if data.iloc[index, 20] not in self.has_fio_doers:
                    row_names.append(data.iloc[index].name)
            for name in row_names:
                operation_id = data.loc[name, 'id']
                mask = data.iloc[:, 21].apply(lambda x: operation_id in x)
                data[mask].iloc[:, 21].apply(lambda x: x.remove(operation_id))
                mask = data.iloc[:, 22].apply(lambda x: operation_id in x)
                data[mask].iloc[:, 22].apply(lambda x: x.remove(operation_id))
                data = data.drop(name)
            self.set_table_data(table, data)
        except Exception as ex:
            logging.exception(ex)

    def add_operation_for_detail(self):
        indexes = set(map(lambda x: x.row(), self.table_details.selectedIndexes()))
        data: pd.DataFrame = self.table_process.model().df
        dialog = OperationChoiceWindow(self)
        dialog.exec()
        for index in sorted(indexes):
            detail = self.table_details.model().df.loc[index]
            try:
                for operation in dialog.operations:
                    row = len(data)
                    prev_operation_id = data[
                        (data.iloc[:, 15] == detail.iloc[0]) &
                        (data.iloc[:, 16] == detail.iloc[1]) &
                        (data.iloc[:, 17] == detail.iloc[2]) &
                        (data.iloc[:, 18] == detail.iloc[3]) &
                        (data.iloc[:, 19] == detail.iloc[4])
                        ]['id'].max()
                    data = self.add_empty_row(self.table_process, row)
                    data.loc[row, :COLUMNS[15]] = operation
                    data.loc[row, COLUMNS[15]:COLUMNS[19]] = detail
                    if not math.isnan(prev_operation_id):
                        data.loc[row, COLUMNS[22]].append(prev_operation_id)
                        next_operation_id = data.loc[row, 'id']
                        data.loc[data['id'] == prev_operation_id][COLUMNS[21]].iloc[0].append(next_operation_id)

            except Exception as ex:
                logging.exception(ex)
        self.set_table_data(self.table_process, data)

    def add_operation(self):
        try:
            data = self.table_process.model().df
            index = self.table_process.currentIndex().row()
            dialog = OperationChoiceWindow(self)
            dialog.exec()
            if index == -1:
                current_row = len(data)
            else:
                current_row = data.iloc[index].name

            for i in range(len(dialog.operations)):
                new_name = current_row + float(f'0.{i + 1}')
                data = self.add_empty_row(self.table_process, new_name)
                data.loc[new_name, :COLUMNS[15]] = dialog.operations[i]
            self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def open_select_columns_dialog(self):
        dialog = SelectColumnsWindow(self)
        dialog.exec()
        self.displayed_columns = dialog.selected_columns
        for i, value in enumerate(COLUMNS):
            self.table_process.setColumnHidden(i, value not in self.displayed_columns)

    def double_click_process_table(self, index):
        try:
            if index.column() in [21, 22]:
                data = self.table_process.model().df
                current_ids = copy(data.iloc[index.row(), index.column()])
                current_id = data.iloc[index.row(), 20]
                filtered_data = data[data.iloc[:, 20] != data.iloc[index.row(), 20]]
                if index.column() == 21:
                    logging.debug('Диалоговое окно выбора следующей операции')
                    column = 22
                elif index.column() == 22:
                    logging.debug('Диалоговое окно выбора предыдущей операции')
                    column = 21
                dialog = SelectNextPrevOperationWindow(self, filtered_data, current_ids)
                dialog.exec()
                data.iloc[index.row(), index.column()].clear()
                data.iloc[index.row(), index.column()].extend(dialog.operations_ids)
                for operation_id in dialog.operations_ids:
                    if current_id not in data.loc[data['id'] == operation_id].iloc[0, column]:
                        data.loc[data['id'] == operation_id].iloc[0, column].append(current_id)
                for operation_id in current_ids:
                    if operation_id not in dialog.operations_ids:
                        data.loc[data['id'] == operation_id].iloc[0, column].remove(current_id)
                self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def draw_graph(self):
        draw_graph(self.table_process.model().df)

    def search(self):
        try:
            if self.search_data is None:
                self.search_data = self.table_process.model().df
            data = self.search_data
            data = data[
                data.apply(lambda x: x.astype(str).str.contains(self.searchbar.text(), case=False, regex=False)).any(
                    axis=1)]
            # for i in range(len(self.table_process.model().df)):
            #     self.table_process.setRowHidden(i, True)
            # for i in range(len(data)):
            #     self.table_process.setRowHidden(int(data.iloc[i].name), False)
            self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def get_orders_models(self):
        try:
            logging.debug('Начало запроса заказ-моделей')
            url = "http://127.0.0.1:8000/tehnolog/orders_models"
            response = requests.get(url)
            self.orders_models = json.loads(response.content)
            self.main_window.statusBar().clearMessage()
        except requests.exceptions.ConnectionError as ex:
            logging.exception(ex)
            self.not_connection()
            self.main_window.statusBar().showMessage("⛔Нет связи с сервером!")
        except Exception as ex:
            logging.exception(ex)
            self.not_connection()
            self.main_window.statusBar().showMessage("Ошибка получения данных!")

    def not_connection(self):
        try:
            self.orders_models = []
            self.model_edit.setDisabled(True)
            self.model_edit.clear()
            self.table_process.setDisabled(True)
        except Exception as ex:
            logging.exception(ex)

    def on_order_model_click(self):
        try:
            self.order_model_select.clear()
            self.order_model_select.addItem("Не выбрано")
            self.get_orders_models()
            # thread = threading.Thread(target=self.get_orders_models)
            # thread.start()
            # thread.join()
            logging.debug('Завершение запроса заказ-моделей')
            for order_model in self.orders_models:
                self.order_model_select.addItem(order_model['order_model'])
        except Exception as ex:
            logging.exception(ex)

    def save_process_to_file(self, upload=False):
        try:
            order = self.order_model_select.currentText().split('_')[0]
            model = self.model_edit.text()
            order_model = f'{order}_{model}'

            if len(self.searchbar.text()) != 0:
                logging.debug('Примененный фильтр сброшен перед сохранением')
                data = self.search_data
            else:
                data = self.table_process.model().df

            if data.empty:
                logging.debug('Нет данных. Процесс не сохранен')
                self.main_window.statusBar().showMessage("⛔Ошибка сохранения! Процесс пустой!")
                return

            if not os.path.exists('data'):
                os.mkdir('data')

            if not os.path.exists(r'data\uploaded'):
                os.mkdir(r'data\uploaded')

            if upload:
                logging.debug(fr'Сохранен файл data\uploaded\{order_model}.json')
                data.to_json(fr'data\uploaded\{order_model}.json', orient='records')
            else:
                logging.debug(fr'Сохранен файл data\{order_model}.json')
                data.to_json(fr'data\{order_model}.json', orient='records')
            self.main_window.statusBar().showMessage("💾Сохранено!")
        except Exception as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("⛔Ошибка сохранения!")

    def save_to_xlsx(self):
        try:
            order = self.order_model_select.currentText().split('_')[0]
            model = self.model_edit.text()
            order_model = f'{order}_{model}'
            name = QFileDialog.getSaveFileName(self, 'Сохранить файл', f"{order_model}.xlsx")[0]
            if name:
                shutil.copyfile('template.xlsx', name)
                new_rows_count = len(self.table_process.model().df)
                wb = load_workbook(name)
                ws = wb['Sheet1']
                ws.title = model
                ws['B2'] = f"Расценки на работу по котлу  {model}"
                ws.insert_rows(5, amount=new_rows_count)
                ws[f'L{5 + new_rows_count}'] = f'=SUM(L5:L{4 + new_rows_count})'
                ws[f'M{5 + new_rows_count}'] = f'=SUM(M5:M{4 + new_rows_count})'
                ws[f'O{5 + new_rows_count}'] = f'=SUM(O5:O{4 + new_rows_count})'
                ws[f'L{7 + new_rows_count}'] = f'=L{5 + new_rows_count}/G2'
                wb.save(name)
                with pd.ExcelWriter(name, if_sheet_exists='overlay', mode='a') as writer:
                    self.table_process.model().df.to_excel(
                        excel_writer=writer,
                        sheet_name=model,
                        index=False,
                        header=False,
                        float_format="%.2f",
                        columns=COLUMNS[:15],
                        startrow=4,
                    )
                logging.debug(f'Процесс сохранен в файл Excel {name}')
        except Exception as ex:
            logging.exception(ex)

    def open_process_from_file(self, order_model):
        try:
            if self.process_is_upload and os.path.exists(fr'data\uploaded\{order_model}.json'):
                file = fr'data\uploaded\{order_model}.json'
                data = pd.read_json(file, orient='records', dtype={COLUMNS[0]: 'str', COLUMNS[4]: 'str'})
                logging.debug(fr'Открыт файл процесса {file}')
            elif os.path.exists(fr'data\{order_model}.json'):
                file = fr'data\{order_model}.json'
                data = pd.read_json(file, orient='records', dtype={COLUMNS[0]: 'str', COLUMNS[4]: 'str'})
                logging.debug(fr'Открыт файл процесса {file}')
            else:
                logging.debug('Новый процесс')
                data = self.initial_data

            max_id = data[COLUMNS[20]].max()
            self.table_process.model().last_id = max_id if pd.notna(max_id) else 0
            self.set_table_data(self.table_process, data)
        except Exception as ex:
            logging.exception(ex)

    def on_order_model_select(self, item):
        try:
            selected_order_model = self.order_model_select.itemText(item)
            for order_model in self.orders_models:
                if selected_order_model == order_model['order_model']:
                    self.model_edit.setText(order_model['model'])
                    self.is_planned = order_model['order_status'] != 'не запланировано'
                    self.process_is_upload = order_model['td_status'] == 'утверждено' and not order_model['on_change']
                    self.has_fio_doers = order_model['has_fio_doers']
                    self.model_edit.setDisabled(self.is_planned)
                    self.table_process.setDisabled(False)
                    if self.process_is_upload:
                        self.change_button.show()
                        self.is_on_change = False
                    else:
                        self.change_button.hide()
                        self.is_on_change = True
                    self._create_process_table_context_menu()
                    self.open_process_from_file(order_model['order_model'])
                    self.process_status_label.setText(
                        f'Статус: {"загружен" if self.process_is_upload else "не загружен"}, '
                        f'{order_model["order_status"]}, '
                        f'{"распределено" if self.has_fio_doers else "не распределено"}'
                    )
                    break
                else:
                    self.model_edit.clear()
                    self.model_edit.setDisabled(True)
                    self.table_process.setDisabled(True)
                    self.process_status_label.setText('Статус: ')
        except Exception as ex:
            logging.exception(ex)

    def set_change_status(self, data):
        self.main_window.statusBar().showMessage("Изменение статуса...")
        try:
            url = "http://127.0.0.1:8000/tehnolog/change_st_status"
            payload = json.dumps(data)
            response = requests.post(url, data=payload)
            data = json.loads(response.content)
            logging.debug(data['message'])
            if response.status_code == 200:
                order = self.order_model_select.currentText().split('_')[0]
                self.on_order_model_click()
                self.order_model_select.setCurrentText(f'{order}_{self.model_edit.text()}')
                self.on_order_model_select(self.order_model_select.currentIndex())
            self.main_window.statusBar().showMessage(data['message'])
        except requests.exceptions.ConnectionError as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("⛔Нет связи с сервером!")
        except Exception as ex:
            logging.exception(ex)
            self.main_window.statusBar().showMessage("Ошибка отправки запроса!")

    def take_on_change(self):
        try:
            message = (f'Вы уверены, что хотите откорректировать процесс?\n'
                       f'При корректировке сменные задания будут недоступны\n'
                       f'для планирования и распределения.\n'
                       f'Для возможности дальнейшего распределения необходимо\n'
                       f'обязательно повторно загрузить процесс!')
            answer = QMessageBox.critical(
                self,
                'Предупреждение!',
                message,
                buttons=QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
            if answer == QMessageBox.StandardButton.Yes:
                ids = self.table_process.model().df[COLUMNS[20]].tolist()
                data = {
                    'model_order_query': self.order_model_select.currentText(),
                    'tech_ids': list(set(ids) - set(self.has_fio_doers)),
                    'status': 'корректировка'
                }
                self.set_change_status(data)
                # thread = threading.Thread(target=self.set_change_status, args=(data,))
                # thread.start()
        except Exception as ex:
            logging.exception(ex)


class SelectXlsxSheetWindow(QDialog):

    def __init__(self, parent=None, sheets=None):
        super(SelectXlsxSheetWindow, self).__init__(parent)
        self.sheets = sheets
        self.selected_sheet = ''
        layout = QGridLayout()
        layout.setSpacing(10)

        self.setLayout(layout)

        ok_button = QPushButton(text="OK", parent=self)
        ok_button.clicked.connect(self.on_confirm_sheet_select)

        label = QLabel()
        label.setText('Выберите модель:')
        self.sheet_select = QComboBox()
        self.sheet_select.addItem("Не выбрано")
        self.sheet_select.addItems(self.sheets)
        self.sheet_select.setMinimumWidth(200)

        layout.addWidget(label, 0, 0, 1, 1, Qt.AlignmentFlag.AlignLeft)
        layout.addWidget(self.sheet_select, 0, 1, 1, 1, Qt.AlignmentFlag.AlignLeft)
        layout.addWidget(ok_button, 1, 0, 1, 2, Qt.AlignmentFlag.AlignHCenter)

    def on_confirm_sheet_select(self):
        index = self.sheet_select.currentIndex()
        if index > 0:
            self.selected_sheet = self.sheet_select.itemText(index)
        else:
            self.selected_sheet = ''
        self.close()


class OperationChoiceWindow(QDialog):
    def __init__(self, parent=None):
        super(OperationChoiceWindow, self).__init__(parent)
        self.setWindowTitle('Выбор операции')
        layout = QGridLayout()
        layout.setSpacing(10)
        self.operations = []
        self.data = None

        self.table_operations = QTableView()
        self.table_operations.setGeometry(0, 0, 800, 600)

        data = pd.read_excel(r'operations.xlsx', index_col=0)
        model = PandasModel(data, 0)

        self.table_operations.setModel(model)

        self.setMinimumWidth(self.table_operations.width() * 2)
        self.setMinimumHeight(self.table_operations.height())

        self.searchbar = QLineEdit()
        self.searchbar.textChanged.connect(self.search)

        from_file_button = QPushButton(text="Загрузить операции из файла", parent=self)
        from_file_button.clicked.connect(self.open_from_file)

        self.setLayout(layout)

        ok_button = QPushButton(text="OK", parent=self)
        ok_button.clicked.connect(self.on_confirm_operation_select)

        layout.addWidget(self.searchbar, 0, 0, 1, 2)
        layout.addWidget(from_file_button, 0, 2, 1, 1)
        layout.addWidget(self.table_operations, 1, 0, 1, 20)
        layout.addWidget(ok_button, 2, 0, 1, 2)

    def set_table_data(self, table, data):
        try:
            model = PandasModel(data, 0)
            table.setModel(model)
        except Exception as ex:
            logging.exception(ex)

    def open_from_file(self):
        xlsx_file_name = QFileDialog.getOpenFileName(
            self,
            caption="Выберите файл",
            directory=r"D:\\",
            filter="Excel Files (*.xlsx);;",
        )[0]
        if xlsx_file_name:
            get_all_operations(xlsx_file_name)
            data = pd.read_excel(r'operations.xlsx', index_col=0)
            self.set_table_data(self.table_operations, data)

    def search(self):
        try:
            if self.data is None:
                self.data = self.table_operations.model().df
            data = self.data
            data = data[
                data.apply(lambda x: x.astype(str).str.contains(self.searchbar.text(), case=False, regex=False)).any(
                    axis=1)]
            self.set_table_data(self.table_operations, data)
        except Exception as ex:
            logging.exception(ex)

    def on_confirm_operation_select(self):
        indexes = set(map(lambda x: x.row(), self.table_operations.selectedIndexes()))
        data = self.table_operations.model().df
        for index in indexes:
            self.operations.append(data.iloc[index])
        self.close()


class SelectColumnsWindow(QDialog):
    def __init__(self, parent=None):
        super(SelectColumnsWindow, self).__init__(parent)
        self.setWindowTitle('Выбор отображаемых столбцов')
        layout = QGridLayout()
        layout.setSpacing(5)
        self.selected_columns = parent.displayed_columns

        self.setLayout(layout)

        self.checkboxes = []
        for column in COLUMNS:
            checkbox = QCheckBox(column)
            checkbox.setChecked(column in self.selected_columns)
            self.checkboxes.append(checkbox)

        ok_button = QPushButton(text="OK", parent=self)
        ok_button.clicked.connect(self.on_confirm_columns_select)

        select_all_button = QPushButton(text="", parent=self)
        select_all_button.setIcon(QIcon("img/select-all.svg"))
        select_all_button.clicked.connect(self.select_all)

        select_none_button = QPushButton(text="", parent=self)
        select_none_button.setIcon(QIcon("img/select-none.svg"))
        select_none_button.clicked.connect(self.select_none)

        layout.addWidget(select_all_button, 0, 0, 1, 1, Qt.AlignmentFlag.AlignLeft)
        layout.addWidget(select_none_button, 0, 1, 1, 1, Qt.AlignmentFlag.AlignLeft)

        last = 0

        for i, checkbox in enumerate(self.checkboxes):
            layout.addWidget(checkbox, 1 + i, 0, 1, 11)
            last = 1 + i

        layout.addWidget(ok_button, last + 1, 2, 1, 6)

    def on_confirm_columns_select(self):
        self.selected_columns = [checkbox.text() for checkbox in self.checkboxes if checkbox.isChecked()]
        self.close()

    def select_all(self):
        for checkbox in self.checkboxes:
            checkbox.setChecked(True)

    def select_none(self):
        for checkbox in self.checkboxes:
            checkbox.setChecked(False)


class SelectNextPrevOperationWindow(OperationChoiceWindow):
    def __init__(self, parent=None, data=None, current_ids=None):
        super(OperationChoiceWindow, self).__init__(parent)
        self.setWindowTitle('Выбор операции')
        layout = QGridLayout()
        layout.setSpacing(10)
        self.operations_ids = current_ids
        self.data = data

        self.table_operations = QTableView()
        self.table_operations.setGeometry(0, 0, 800, 600)

        model = PandasModel(data, 0)

        self.table_operations.setModel(model)

        self.setMinimumWidth(self.table_operations.width() * 2)
        self.setMinimumHeight(self.table_operations.height())

        self.searchbar = QLineEdit()
        self.searchbar.textChanged.connect(self.search)

        self.setLayout(layout)

        ok_button = QPushButton(text="OK", parent=self)
        ok_button.clicked.connect(self.on_confirm_operation_select)

        layout.addWidget(self.searchbar, 0, 0, 1, 2)
        layout.addWidget(self.table_operations, 1, 0, 1, 20)
        layout.addWidget(ok_button, 2, 0, 1, 2)

    def on_confirm_operation_select(self):
        indexes = set(map(lambda x: x.row(), self.table_operations.selectedIndexes()))
        data = self.table_operations.model().df
        self.operations_ids = []
        for index in indexes:
            self.operations_ids.append(data.iloc[index, 20])
        self.close()


if __name__ == "__main__":
    app = QApplication([])
    with open('light.css', 'r') as style_file:
        theme = style_file.read()
    app.setStyleSheet(theme)
    window = Window()
    window.show()
    try:
        sys.exit(app.exec())
    except Exception as ex:
        logging.exception(ex)
